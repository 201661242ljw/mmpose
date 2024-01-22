import os.path
import cv2
import shutil
import openpyxl
import colorsys
import copy
import json
import os

import numpy as np


def get_finale_ann():
    src_dir = r"../data/00_Tower_Dataset"

    json_path = r"../data/coco/annotations/person_keypoints_train2017.json"
    template_data = json.load(open(json_path, "r", encoding="utf-8"), strict=False)

    for input_size in [256, 384, 512, 640, 768, 1024, 1280]:
        img_id = 0
        ann_id = 0
        for dataset in ['train', 'val', 'test']:
            output_data = copy.deepcopy(template_data)
            output_data['categories'] = [{'supercategory': "tower",
                                          'id': 0,
                                          'name': 'tower',
                                          'keypoints': ['edgepoint',
                                                        'skeletonpoints',
                                                        'edgepoint_2',
                                                        'bileipoint',
                                                        'skeletonpoint_2'],
                                          'skeleton': []
                                          }]
            output_data['images'] = []
            output_data['annotations'] = []

            json_path = r"{}/{}/anns/tower_info_{}.json".format(src_dir, input_size, dataset)
            data = json.load(open(json_path, "r", encoding="utf-8"), strict=False)

            for (img_name, temp_dict) in data.items():
                width = temp_dict['width']
                height = temp_dict['height']
                keypoint_list = [width // 2, height // 2, 2] * 5
                [x1, y1, x2, y2] = temp_dict['bbox']

                img_dict = {
                    "license": 1,
                    "file_name": img_name,
                    "coco_url": "",
                    "height": height,
                    "width": width,
                    "date_captured": "",
                    "flickr_url": "",
                    "id": img_id
                }

                ann_dict = {
                    "segmentation": [[x1, y1, x2, y1, x2, y2, x1, y2]],
                    "true_points_1": np.array(temp_dict["new_keypoints_1"])[:, 1:].astype(np.int32).tolist(),
                    "true_skeletons_1": temp_dict["new_skeletons_1"],
                    "true_points_2": np.array(temp_dict["new_keypoints_2"])[:, 1:].astype(np.int32).tolist(),
                    "true_skeletons_2": temp_dict["new_skeletons_2"],
                    "true_points_3": np.array(temp_dict["new_keypoints_3"])[:, 1:].astype(np.int32).tolist(),
                    "true_skeletons_3": temp_dict["new_skeletons_3"],
                    "num_keypoints": 5,
                    "area": int((x2 - x1) * (y2 - y1)),
                    "iscrowd": 0,
                    "keypoints": keypoint_list,
                    "image_id": img_id,
                    "bbox": [x1, y1, x2 - x1, y2 - y1],
                    "category_id": 1,
                    "id": ann_id
                }
                output_data['images'].append(img_dict)
                output_data['annotations'].append(ann_dict)

                img_id += 1
                ann_id += 1

            out_json_path = r"{}/{}/anns/tower_keypoints_{}_2.json".format(src_dir, input_size, dataset)
            f = open(out_json_path, "w", encoding="utf-8")
            f.write(json.dumps(output_data, ensure_ascii=False, indent=4))
            f.close()


def main():
    # src_dir_img = r"E:\LJW\mmpose_\00_LJW\tower_dataset_12456_train_val_test\imgs"
    num = 0

    # ue_json_dir = r"E:\LJW\UE4\UAV\datasets\json"
    ue_3_json_dir = r"../data/UE4/3_json"
    ue_3_img_dir = r"../data/UE4\3_img"
    ue_3_img_list = os.listdir(ue_3_img_dir)
    ue_3_img_num = len(ue_3_img_list)
    rand = np.random.rand(ue_3_img_num)
    train_img_list_3 = np.array(ue_3_img_list)[rand < 0.7].tolist()
    val_img_list_3 = np.array(ue_3_img_list)[np.logical_and(rand < 0.8, rand >= 0.7)].tolist()
    test_img_list_3 = np.array(ue_3_img_list)[rand >= 0.8].tolist()

    ue_7_json_dir = r"../data/UE4/7_json"
    ue_7_img_dir = r"../data/UE4\7_img"
    ue_7_img_list = os.listdir(ue_7_img_dir)
    ue_7_img_num = len(ue_7_img_list)
    rand = np.random.rand(ue_7_img_num)
    train_img_list_7 = np.array(ue_7_img_list)[rand < 0.7].tolist()
    val_img_list_7 = np.array(ue_7_img_list)[np.logical_and(rand < 0.8, rand >= 0.7)].tolist()
    test_img_list_7 = np.array(ue_7_img_list)[rand >= 0.8].tolist()

    json_dir = r"../data/tower_dataset_12456_train_val_test/annotations"
    img_dir = r"../data/tower_dataset_12456_train_val_test/imgs"

    json_partial_dir = r"../data/partial_anns"
    img_partial_dir = r"../data/partial_all_imgs"

    for dataset_idx, dataset in enumerate(['train', 'val', 'test']):
        temp_data = {}

        json_path = os.path.join(json_dir, "0_keypoints_{}.json".format(dataset))
        data = json.load(open(json_path, 'r', encoding='utf-8'), strict=False)
        keypoint_name_list = np.array(data['categories'][0]["keypoint"])
        images, annotations = data["images"], data["annotations"]

        json_path_partial = os.path.join(json_partial_dir, "0_keypoints_{}.json".format(dataset))
        data_partial = json.load(open(json_path_partial, 'r', encoding='utf-8'), strict=False)
        images_partial, annotations_partial = data_partial["images"], data_partial["annotations"]

        for sheet_order in range(1, 4):

            need_flip_list = []

            points_data = {}
            book = openpyxl.load_workbook(r"new_points.xlsx")
            sh = book[f'Sheet{sheet_order * 2}']
            row_num = 2
            while sh[f"A{row_num}"].value is not None:
                new_name = sh[f"A{row_num}"].value
                type = sh[f"B{row_num}"].value
                points_data[new_name] = {"type": type,
                                         "points": []}
                c_num = 67
                while sh[f"{chr(c_num)}{row_num}"].value is not None:
                    points_data[new_name]["points"].append(sh[f"{chr(c_num)}{row_num}"].value)
                    c_num += 1
                row_num += 1
            b = json.dumps(points_data, ensure_ascii=False, indent=4)
            f2 = open(r"new_points.json", 'w', encoding='utf-8')
            f2.write(b)
            f2.close()

            skeleton_data = {}
            sh = book[f'Sheet{sheet_order * 2 + 1}']
            row_num = 2
            while sh[f"A{row_num}"].value is not None:
                p1 = sh[f"A{row_num}"].value
                p2 = sh[f"B{row_num}"].value
                t = sh[f"C{row_num}"].value
                skeleton_data[f"{row_num - 1}"] = {"p1": p1,
                                                   "p2": p2,
                                                   "type": t}
                row_num += 1
            b = json.dumps(skeleton_data, ensure_ascii=False, indent=4)
            f2 = open(r"new_skeletons.json", 'w', encoding='utf-8')
            f2.write(b)
            f2.close()

            if True:
                # ------------------------------------------------------------------------------------------------------
                # true imgs
                # ------------------------------------------------------------------------------------------------------
                for image, annotation in zip(images, annotations):
                    assert image["id"] == annotation["image_id"]
                    keypoints = np.array(annotation["keypoints"]).reshape(-1, 3)
                    keypoints, real_keypoint_name_list = keypoints[keypoints[:, -1] != 0], keypoint_name_list[
                        keypoints[:, -1] != 0]
                    image_name = image["file_name"]
                    if not image_name in temp_data.keys():
                        temp_data[image_name] = {}

                    temp_data[image_name]["file_path"] = os.path.join(img_dir, image_name)
                    temp_data[image_name]['width'] = image['width']
                    temp_data[image_name]['height'] = image['height']
                    temp_data[image_name]['old_keypoints'] = np.hstack(
                        [np.expand_dims(real_keypoint_name_list, axis=1), keypoints[:, :2]]).tolist()
                    temp_data[image_name][f'new_keypoints_{sheet_order}'] = []
                    temp_data[image_name][f'new_skeletons_{sheet_order}'] = []
                    temp_data[image_name]['if_ue'] = False
                    num += 1
                    real_keypoint_name_list = real_keypoint_name_list.tolist()
                    print(num)

                    new_point_names = []
                    new_xs = []
                    new_ys = []
                    new_types = []

                    special_flag = 0
                    special_flag_2 = 0
                    need_flip = False

                    if real_keypoint_name_list[0][0] == "4":
                        if keypoints[real_keypoint_name_list.index("4_1_1")][0] > \
                                keypoints[real_keypoint_name_list.index("4_1_4")][0]:
                            need_flip_list.append(image_name.split(".")[0])
                            need_flip = True

                    for (new_point_name, temp_dict) in points_data.items():
                        xs = []
                        ys = []

                        for p_name in temp_dict["points"]:
                            if p_name in real_keypoint_name_list:
                                xs.append(keypoints[real_keypoint_name_list.index(p_name)][0])
                                ys.append(keypoints[real_keypoint_name_list.index(p_name)][1])

                        aaa = {
                            '1': "3",
                            "2": "4",
                            "3": "1",
                            '4': "2",
                            "5": "7",
                            '6': "8",
                            "7": "5",
                            '8': "6",
                        }

                        if len(xs) != 0:
                            x = int(np.average(np.array(xs)))
                            y = int(np.average(np.array(ys)))
                            if need_flip and sheet_order == 3:
                                [tower_type, floor, _, num__] = new_point_name.split("_")
                                if floor != "4":
                                    new_point_name = f"{tower_type}_{floor}_{_}_{aaa[num__]}"

                            new_point_names.append(new_point_name)
                            new_xs.append(x)
                            new_ys.append(y)
                            new_type = points_data[new_point_name]["type"]
                            new_types.append(new_type)
                            temp_data[image_name][f'new_keypoints_{sheet_order}'].append(
                                [new_point_name, x, y, new_type])
                            if sheet_order == 2:
                                if new_point_name == "4_2_4":
                                    special_flag = 1
                            elif sheet_order == 3:
                                if new_point_name == "4_2_sk_1":
                                    special_flag = 1
                                if new_point_name == "4_2_e_1":
                                    special_flag_2 = 1

                    if (special_flag_2 == 0 and special_flag == 1) and sheet_order == 3:
                        for [new_point_name, x, y, t_] in temp_data[image_name][f'new_keypoints_{sheet_order}']:
                            if new_point_name == "4_2_sk_1":
                                new_point_names.append("4_2_sk_5")
                                new_xs.append(x)
                                new_ys.append(y)
                                new_types.append(4)
                                temp_data[image_name][f'new_keypoints_{sheet_order}'].append(
                                    ["4_2_sk_5", x, y, 4])
                            if new_point_name == "4_2_sk_2":
                                new_point_names.append("4_2_sk_6")
                                new_xs.append(x)
                                new_ys.append(y)
                                new_types.append(4)
                                temp_data[image_name][f'new_keypoints_{sheet_order}'].append(
                                    ["4_2_sk_6", x, y, 4])
                            if new_point_name == "4_2_sk_3":
                                new_point_names.append("4_2_sk_7")
                                new_xs.append(x)
                                new_ys.append(y)
                                new_types.append(6)
                                temp_data[image_name][f'new_keypoints_{sheet_order}'].append(
                                    ["4_2_sk_7", x, y, 6])
                            if new_point_name == "4_2_sk_4":
                                new_point_names.append("4_2_sk_8")
                                new_xs.append(x)
                                new_ys.append(y)
                                new_types.append(6)
                                temp_data[image_name][f'new_keypoints_{sheet_order}'].append(
                                    ["4_2_sk_8", x, y, 6])
                    # skeleton
                    for (_, temp_dict) in skeleton_data.items():
                        p1 = temp_dict['p1']
                        p2 = temp_dict["p2"]
                        t = temp_dict["type"]
                        if special_flag and sheet_order == 2:
                            if p1 == "4_2_2" and p2 == "4_3_2":
                                continue
                        if need_flip and sheet_order == 3:
                            if p1 == "4_3_sk_8" and p2 == "4_4_sk_8":
                                p1 = "4_3_sk_7"
                                p2 = "4_4_sk_7"
                            if p1 == "4_3_sk_5" and p2 == "4_4_sk_7":
                                p1 = "4_3_sk_6"
                                p2 = "4_4_sk_8"

                        if p1 in new_point_names and p2 in new_point_names:
                            x1 = new_xs[new_point_names.index(p1)]
                            y1 = new_ys[new_point_names.index(p1)]
                            t1 = new_types[new_point_names.index(p1)]
                            x2 = new_xs[new_point_names.index(p2)]
                            y2 = new_ys[new_point_names.index(p2)]
                            t2 = new_types[new_point_names.index(p2)]
                            temp_data[image_name][f'new_skeletons_{sheet_order}'].append([x1, y1, t1, x2, y2, t2, t])

                # ------------------------------------------------------------------------------------------------------
                # partial true imgs
                # ------------------------------------------------------------------------------------------------------
                if True:
                    # if not True:
                    for image, annotation in zip(images_partial, annotations_partial):
                        assert image["id"] == annotation["image_id"]
                        keypoints = np.array(annotation["keypoints"]).reshape(-1, 3)
                        keypoints, real_keypoint_name_list = keypoints[keypoints[:, -1] != 0], keypoint_name_list[
                            keypoints[:, -1] != 0]
                        image_name = image["file_name"]
                        part_id = image_name.split(".")[0].split("_")[-1]
                        if len(part_id) == 1:
                            continue
                        if not image_name in temp_data.keys():
                            temp_data[image_name] = {}

                        temp_data[image_name]["file_path"] = os.path.join(img_dir, image_name)
                        temp_data[image_name]['width'] = image['width']
                        temp_data[image_name]['height'] = image['height']
                        temp_data[image_name]['old_keypoints'] = np.hstack(
                            [np.expand_dims(real_keypoint_name_list, axis=1), keypoints[:, :2]]).tolist()
                        temp_data[image_name][f'new_keypoints_{sheet_order}'] = []
                        temp_data[image_name][f'new_skeletons_{sheet_order}'] = []
                        temp_data[image_name]['if_ue'] = False
                        num += 1
                        real_keypoint_name_list = real_keypoint_name_list.tolist()
                        print(num)

                        new_point_names = []
                        new_xs = []
                        new_ys = []
                        new_types = []

                        special_flag = 0
                        special_flag_2 = 0
                        need_flip = False

                        if real_keypoint_name_list[0][0] == "4":
                            base_name = image_name.split("_partial")[0]
                            if base_name in need_flip_list:
                                need_flip = True

                        for (new_point_name, temp_dict) in points_data.items():
                            xs = []
                            ys = []

                            for p_name in temp_dict["points"]:
                                if p_name in real_keypoint_name_list:
                                    xs.append(keypoints[real_keypoint_name_list.index(p_name)][0])
                                    ys.append(keypoints[real_keypoint_name_list.index(p_name)][1])

                            aaa = {
                                '1': "3",
                                "2": "4",
                                "3": "1",
                                '4': "2",
                                "5": "7",
                                '6': "8",
                                "7": "5",
                                '8': "6",
                            }

                            if len(xs) != 0:
                                x = int(np.average(np.array(xs)))
                                y = int(np.average(np.array(ys)))
                                if need_flip and sheet_order == 3:
                                    [tower_type, floor, _, num__] = new_point_name.split("_")
                                    if floor != "4":
                                        new_point_name = f"{tower_type}_{floor}_{_}_{aaa[num__]}"

                                new_point_names.append(new_point_name)
                                new_xs.append(x)
                                new_ys.append(y)
                                new_type = points_data[new_point_name]["type"]
                                new_types.append(new_type)
                                temp_data[image_name][f'new_keypoints_{sheet_order}'].append(
                                    [new_point_name, x, y, new_type])
                                if sheet_order == 2:
                                    if new_point_name == "4_2_4":
                                        special_flag = 1
                                elif sheet_order == 3:
                                    if new_point_name == "4_2_sk_1":
                                        special_flag = 1
                                    if new_point_name == "4_2_e_1":
                                        special_flag_2 = 1

                        if (special_flag_2 == 0 and special_flag == 1) and sheet_order == 3:
                            for [new_point_name, x, y, t_] in temp_data[image_name][f'new_keypoints_{sheet_order}']:
                                if new_point_name == "4_2_sk_1":
                                    new_point_names.append("4_2_sk_5")
                                    new_xs.append(x)
                                    new_ys.append(y)
                                    new_types.append(4)
                                    temp_data[image_name][f'new_keypoints_{sheet_order}'].append(
                                        ["4_2_sk_5", x, y, 4])
                                if new_point_name == "4_2_sk_2":
                                    new_point_names.append("4_2_sk_6")
                                    new_xs.append(x)
                                    new_ys.append(y)
                                    new_types.append(4)
                                    temp_data[image_name][f'new_keypoints_{sheet_order}'].append(
                                        ["4_2_sk_6", x, y, 4])
                                if new_point_name == "4_2_sk_3":
                                    new_point_names.append("4_2_sk_7")
                                    new_xs.append(x)
                                    new_ys.append(y)
                                    new_types.append(6)
                                    temp_data[image_name][f'new_keypoints_{sheet_order}'].append(
                                        ["4_2_sk_7", x, y, 6])
                                if new_point_name == "4_2_sk_4":
                                    new_point_names.append("4_2_sk_8")
                                    new_xs.append(x)
                                    new_ys.append(y)
                                    new_types.append(6)
                                    temp_data[image_name][f'new_keypoints_{sheet_order}'].append(
                                        ["4_2_sk_8", x, y, 6])
                        # skeleton
                        for (_, temp_dict) in skeleton_data.items():
                            p1 = temp_dict['p1']
                            p2 = temp_dict["p2"]
                            t = temp_dict["type"]
                            if special_flag and sheet_order == 2:
                                if p1 == "4_2_2" and p2 == "4_3_2":
                                    continue
                            if need_flip and sheet_order == 3:
                                if p1 == "4_3_sk_8" and p2 == "4_4_sk_8":
                                    p1 = "4_3_sk_7"
                                    p2 = "4_4_sk_7"
                                if p1 == "4_3_sk_5" and p2 == "4_4_sk_7":
                                    p1 = "4_3_sk_6"
                                    p2 = "4_4_sk_8"

                            if p1 in new_point_names and p2 in new_point_names:
                                x1 = new_xs[new_point_names.index(p1)]
                                y1 = new_ys[new_point_names.index(p1)]
                                t1 = new_types[new_point_names.index(p1)]
                                x2 = new_xs[new_point_names.index(p2)]
                                y2 = new_ys[new_point_names.index(p2)]
                                t2 = new_types[new_point_names.index(p2)]
                                temp_data[image_name][f'new_skeletons_{sheet_order}'].append(
                                    [x1, y1, t1, x2, y2, t2, t])

                # ------------------------------------------------------------------------------------------------------
                # UE_3 imgs
                # ------------------------------------------------------------------------------------------------------
                ue_3_img_dataset = [train_img_list_3, val_img_list_3, test_img_list_3][dataset_idx]
                real_keypoint_name_list = []
                for ue_3_img_name in ue_3_img_dataset:
                    num += 1
                    print(num)
                    ue_3_img_path = os.path.join(r"../data/UE4/IMG_", ue_3_img_name)

                    ue_3_img = cv2.imread(ue_3_img_path, 1) // 4
                    height = ue_3_img.shape[0]
                    width = ue_3_img.shape[1]

                    ue_ann_path = os.path.join(ue_3_json_dir, ue_3_img_name.replace("png", "json"))

                    ue_data = json.load(open(ue_ann_path, "r", encoding="utf-8"), strict=False)["points"]
                    if not ue_3_img_name in temp_data.keys():
                        temp_data[ue_3_img_name] = {}

                    temp_data[ue_3_img_name]["file_path"] = ue_3_img_path
                    temp_data[ue_3_img_name]['width'] = width
                    temp_data[ue_3_img_name]['height'] = height
                    temp_data[ue_3_img_name]['old_keypoints'] = []
                    temp_data[ue_3_img_name][f'new_keypoints_{sheet_order}'] = []
                    temp_data[ue_3_img_name][f'new_skeletons_{sheet_order}'] = []
                    temp_data[ue_3_img_name]['if_ue'] = True
                    new_point_names = []
                    new_xs = []
                    new_ys = []
                    new_types = []
                    for p_name in ue_data.keys():
                        temp_data[ue_3_img_name]['old_keypoints'].append(
                            [p_name, ue_data[p_name]['x'], ue_data[p_name]['y']])
                    for (new_point_name, temp_dict) in points_data.items():
                        xs = []
                        ys = []

                        for p_name in temp_dict["points"]:
                            if p_name in ue_data.keys():
                                xs.append(ue_data[p_name]['x'])
                                ys.append(ue_data[p_name]['y'])

                        if len(xs) != 0:
                            x = int(np.average(np.array(xs)))
                            y = int(np.average(np.array(ys)))
                            if need_flip:
                                [tower_type, floor, _, num] = new_point_name.split("_")
                                num = int(num)
                                if num < 5:
                                    new_point_name = f"{tower_type}_{floor}_e_{int(4 - num)}"
                                else:
                                    new_point_name = f"{tower_type}_{floor}_e_{int(13 - num)}"
                            # for p_name in temp_dict["points"]:
                            #     if p_name in real_keypoint_name_list:
                            #         img = cv2.line(img, pt1=(x, y), pt2=(keypoints[real_keypoint_name_list.index(p_name)][0],
                            #                                              keypoints[real_keypoint_name_list.index(p_name)][1]),
                            #                        color=(255, 255, 255), thickness=2)

                            # text = new_point_name
                            # org = (x, y)
                            # fontFace = cv2.FONT_HERSHEY_SIMPLEX
                            # fontScale = 1
                            # color = [(0, 0, 255), (0, 255, 255), (0, 255, 0), (255, 0, 0), (255, 255, 0)][temp_dict['type'] - 1]
                            # thickness = 2

                            # img = cv2.circle(img, center=(x, y), radius=30, color=color, thickness=-1)
                            # cv2.putText(img, text, org, fontFace, fontScale, color, thickness)
                            new_point_names.append(new_point_name)
                            new_xs.append(x)
                            new_ys.append(y)
                            new_types.append(temp_dict["type"])
                            temp_data[ue_3_img_name][f'new_keypoints_{sheet_order}'].append(
                                [new_point_name, x, y, temp_dict["type"]])

                    for (_, temp_dict) in skeleton_data.items():
                        p1 = temp_dict['p1']
                        p2 = temp_dict["p2"]
                        t = temp_dict["type"]
                        if p1 in new_point_names and p2 in new_point_names:
                            x1 = new_xs[new_point_names.index(p1)]
                            y1 = new_ys[new_point_names.index(p1)]
                            t1 = new_types[new_point_names.index(p1)]
                            x2 = new_xs[new_point_names.index(p2)]
                            y2 = new_ys[new_point_names.index(p2)]
                            t2 = new_types[new_point_names.index(p2)]
                            temp_data[ue_3_img_name][f'new_skeletons_{sheet_order}'].append([x1, y1, t1, x2, y2, t2, t])
                    a = 1

                # ------------------------------------------------------------------------------------------------------
                # UE_3 imgs
                # ------------------------------------------------------------------------------------------------------
                ue_7_img_dataset = [train_img_list_7, val_img_list_7, test_img_list_7][dataset_idx]
                real_keypoint_name_list = []
                for ue_7_img_name in ue_7_img_dataset:
                    num += 1
                    print(num)
                    ue_7_img_path = os.path.join(r"../data/UE4/IMG_", ue_7_img_name)

                    ue_7_img = cv2.imread(ue_7_img_path, 1) // 4
                    height = ue_7_img.shape[0]
                    width = ue_7_img.shape[1]

                    ue_ann_path = os.path.join(ue_7_json_dir, ue_7_img_name.replace("png", "json"))

                    ue_data = json.load(open(ue_ann_path, "r", encoding="utf-8"), strict=False)["points"]
                    if not ue_7_img_name in temp_data.keys():
                        temp_data[ue_7_img_name] = {}

                    temp_data[ue_7_img_name]["file_path"] = ue_7_img_path
                    temp_data[ue_7_img_name]['width'] = width
                    temp_data[ue_7_img_name]['height'] = height
                    temp_data[ue_7_img_name]['old_keypoints'] = []
                    temp_data[ue_7_img_name][f'new_keypoints_{sheet_order}'] = []
                    temp_data[ue_7_img_name][f'new_skeletons_{sheet_order}'] = []
                    temp_data[ue_7_img_name]['if_ue'] = True
                    new_point_names = []
                    new_xs = []
                    new_ys = []
                    new_types = []
                    for p_name in ue_data.keys():
                        temp_data[ue_7_img_name]['old_keypoints'].append(
                            [p_name, ue_data[p_name]['x'], ue_data[p_name]['y']])
                    for (new_point_name, temp_dict) in points_data.items():
                        xs = []
                        ys = []

                        for p_name in temp_dict["points"]:
                            if p_name in ue_data.keys():
                                xs.append(ue_data[p_name]['x'])
                                ys.append(ue_data[p_name]['y'])

                        if len(xs) != 0:
                            x = int(np.average(np.array(xs)))
                            y = int(np.average(np.array(ys)))
                            if need_flip:
                                [tower_type, floor, _, num] = new_point_name.split("_")
                                num = int(num)
                                if num < 5:
                                    new_point_name = f"{tower_type}_{floor}_e_{int(4 - num)}"
                                else:
                                    new_point_name = f"{tower_type}_{floor}_e_{int(13 - num)}"
                            # for p_name in temp_dict["points"]:
                            #     if p_name in real_keypoint_name_list:
                            #         img = cv2.line(img, pt1=(x, y), pt2=(keypoints[real_keypoint_name_list.index(p_name)][0],
                            #                                              keypoints[real_keypoint_name_list.index(p_name)][1]),
                            #                        color=(255, 255, 255), thickness=2)

                            # text = new_point_name
                            # org = (x, y)
                            # fontFace = cv2.FONT_HERSHEY_SIMPLEX
                            # fontScale = 1
                            # color = [(0, 0, 255), (0, 255, 255), (0, 255, 0), (255, 0, 0), (255, 255, 0)][temp_dict['type'] - 1]
                            # thickness = 2

                            # img = cv2.circle(img, center=(x, y), radius=30, color=color, thickness=-1)
                            # cv2.putText(img, text, org, fontFace, fontScale, color, thickness)
                            new_point_names.append(new_point_name)
                            new_xs.append(x)
                            new_ys.append(y)
                            new_types.append(temp_dict["type"])
                            temp_data[ue_7_img_name][f'new_keypoints_{sheet_order}'].append(
                                [new_point_name, x, y, temp_dict["type"]])

                    for (_, temp_dict) in skeleton_data.items():
                        p1 = temp_dict['p1']
                        p2 = temp_dict["p2"]
                        t = temp_dict["type"]
                        if p1 in new_point_names and p2 in new_point_names:
                            x1 = new_xs[new_point_names.index(p1)]
                            y1 = new_ys[new_point_names.index(p1)]
                            t1 = new_types[new_point_names.index(p1)]
                            x2 = new_xs[new_point_names.index(p2)]
                            y2 = new_ys[new_point_names.index(p2)]
                            t2 = new_types[new_point_names.index(p2)]
                            temp_data[ue_7_img_name][f'new_skeletons_{sheet_order}'].append([x1, y1, t1, x2, y2, t2, t])
                    a = 1
                print(
                    "------------------------------------------------------------------------------------------------------")

                #         ue_3_img = cv2.line(ue_3_img, pt1=(x1,y1),pt2=(x2,y2),color=(0,0,255), thickness=5)
                # cv2.imwrite(r"test3\{}".format(ue_3_img_name), ue_3_img)
                # exit()
        b = json.dumps(temp_data, ensure_ascii=False, indent=4)
        f2 = open(r"../data/00_new_dataset/tower_info_{}_2.json".format(dataset), 'w', encoding='utf-8')
        f2.write(b)
        f2.close()


def get_sample():
    json_path = r"../data/00_new_dataset/tower_info_train_2.json"
    data = json.load(open(json_path, "r", encoding="utf-8"), strict=False)

    already_draw = []

    for (img_name, img_data) in data.items():
        kts = img_data['new_keypoints_3']
        tower_type = kts[0][0][0]
        if not tower_type in already_draw:
            already_draw.append(tower_type)
            img_path = img_data['file_path']
            img = cv2.imread(img_path, 1) // 8
            for [kt_name, x, y, kt_type] in kts:
                img = cv2.circle(img, center=(x, y), color=pt_colors[kt_type - 1], thickness=-1, radius=3)
                img = cv2.putText(img, kt_name, (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 1)
                # cv2.putText(image, text, (5, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 255), 2)
            cv2.imwrite("../data/label_kts/{}.jpg".format(tower_type), img)
        # if tower_type == "4":
        #     already_draw.append(tower_type)
        #     img_path = img_data['file_path']
        #     img = cv2.imread(img_path, 1) // 8
        #     for [kt_name, x, y, kt_type] in kts:
        #         img = cv2.circle(img, center=(x, y), color=pt_colors[kt_type - 1], thickness=-1, radius=3)
        #         img = cv2.putText(img, kt_name, (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 255), 2)
        #         # cv2.putText(image, text, (5, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (0, 0, 255), 2)
        #     cv2.imwrite("../data/label_kts/{}.jpg".format(os.path.basename(img_path)), img)


def step2():
    # type_ = 2
    for new_size in [256, 384, 512, 640, 768, 1024, 1280]:
        ann_dir = r"../data/00_new_dataset"
        save_dir = r"../data/00_Tower_Dataset/{}".format(new_size)

        if os.path.exists(os.path.join(save_dir, "imgs_show")):
            shutil.rmtree(os.path.join(save_dir, "imgs_show"))

        os.makedirs(os.path.join(save_dir, "imgs_show"))

        if not os.path.exists(os.path.join(save_dir, "anns")):
            os.makedirs(os.path.join(save_dir, "anns"))
        if not os.path.exists(os.path.join(save_dir, "imgs")):
            os.makedirs(os.path.join(save_dir, "imgs"))

        already_draw_list = []
        num = 0
        for dataset in ["train", "val", "test"]:
            json_path = os.path.join(ann_dir, f"tower_info_{dataset}_2.json")
            data = json.load(open(json_path, "r", encoding="utf-8"), strict=False)
            for (img_name, img_data) in data.items():
                num += 1
                print(new_size, "  \t", num)
                img_path = img_data["file_path"]
                img_width = img_data["width"]
                img_height = img_data["height"]
                # if os.path.exists(save_img_path):
                #     continue
                old_points = np.array(img_data['old_keypoints'])
                first_point_name = old_points[0, 0]

                # if "point" in first_point_name:
                #     flag = "point"
                # else:
                flag = first_point_name[0]

                for type_ in range(1, 4):
                    new_points = np.array(img_data[f"new_keypoints_{type_}"])
                    new_skeletons = np.array(img_data[f"new_skeletons_{type_}"])

                    old_xs = old_points[:, 1].astype(int)
                    old_ys = old_points[:, 2].astype(int)

                    x1 = np.min(old_xs)
                    x2 = np.max(old_xs)
                    width = x2 - x1
                    y1 = np.min(old_ys)
                    y2 = np.max(old_ys)
                    height = y2 - y1

                    x1 = max(0, int(x1 - width * 0.05))
                    x2 = min(img_width - 1, int(x2 + width * 0.05))
                    y1 = max(0, int(y1 - height * 0.05))
                    y2 = min(img_height - 1, int(y2 + height * 0.05))

                    width = x2 - x1
                    height = y2 - y1
                    s = int(max(width, height) * 1.05)

                    new_width = int(img_width / s * new_size)
                    new_height = int(img_height / s * new_size)

                    old_xs = (old_xs / s * new_size).astype(int)
                    old_ys = (old_ys / s * new_size).astype(int)

                    new_points[:, 1:3] = (new_points[:, 1:3].astype(int) / s * new_size).astype(int)

                    if new_skeletons.shape[0] != 0:
                        new_skeletons[:, :2] = (new_skeletons[:, :2] / s * new_size).astype(int)
                        new_skeletons[:, 3:5] = (new_skeletons[:, 3:5] / s * new_size).astype(int)

                    img_data["width"] = new_width
                    img_data["height"] = new_height
                    img_data["bbox"] = [int(x1 / s * new_size), int(y1 / s * new_size), int(x2 / s * new_size),
                                        int(y2 / s * new_size)]
                    img_data['old_keypoints'] = np.vstack([old_points[:, 0], old_xs, old_ys]).T.tolist()
                    img_data[f"new_keypoints_{type_}"] = new_points.tolist()
                    img_data[f"new_skeletons_{type_}"] = new_skeletons.tolist()

                    save_path = os.path.join(save_dir, "imgs", img_name)
                    if not os.path.exists(save_path):
                        img = cv2.imread(img_path, 1)
                        new_img = cv2.resize(img, (new_width, new_height))
                        cv2.imwrite(save_path, new_img)
                    # if flag != "4":
                    if flag in already_draw_list:
                        continue

                    img = cv2.imread(img_path, 1)
                    dst_path = os.path.join(save_dir, "imgs_show", img_name)
                    if not os.path.exists(dst_path):
                        shutil.copy(img_path, dst_path)
                    new_img = cv2.resize(img, (new_width, new_height))

                    new_img = new_img // 3

                    for [p_name, x, y, t] in img_data[f'new_keypoints_{type_}']:
                        x1 = int(x)
                        y1 = int(y)
                        t = int(t)
                        new_img = cv2.circle(new_img, center=(x1, y1), color=pt_colors[t - 1], thickness=-1,
                                             radius=new_size // [64, 80, 96][type_ - 1])

                    for [x1, y1, t1, x2, y2, t2, kt] in img_data[f"new_skeletons_{type_}"]:
                        new_img = cv2.arrowedLine(new_img, pt1=(x1, y1), pt2=((x2 + x1) // 2, (y1 + y2) // 2),
                                                  color=sk_colores[kt - 1], thickness=[8, 5, 2][type_ - 1])
                        new_img = cv2.line(new_img, pt1=(x1, y1), pt2=(x2, y2), color=sk_colores[kt - 1], thickness=[8, 5, 2][type_ - 1])
                    new_img = cv2.rectangle(new_img, pt1=(img_data["bbox"][0], img_data["bbox"][1]),
                                            pt2=(img_data["bbox"][2], img_data["bbox"][2]), color=(0, 255, 0),
                                            thickness=1)

                    save_img_path = os.path.join(save_dir, "imgs_show", img_name.split(".")[0] + f"_{type_}.JPG")
                    cv2.imwrite(save_img_path, new_img)
                already_draw_list.append(flag)
            save_json_path = os.path.join(save_dir, "anns", f"tower_info_{dataset}.json")
            b = json.dumps(data, ensure_ascii=False, indent=4)
            f2 = open(save_json_path, 'w', encoding='utf-8')
            f2.write(b)
            f2.close()


def get_complex_skeletons():
    import openpyxl

    temp_dict = [
        {
            "description": "前层边缘点和下层支撑点",
            "floors": [
                [[11, 11], [12, 12], [13, 13], [21, 21], [22, 22], [31, 31], [32, 32], [33, 33], [41, 41], [44, 44],
                 [42, 42], [71, 71], [72, 72], [73, 73]],
                [[51, 51], [61, 61]]
            ],
            "name_pairs": [
                [["sk_1", "e_1"], ["sk_2", "e_2"]],
                [["sk_1", "e_1"], ["sk_4", "e_2"]]
            ]
        },
        {
            "description": "后层边缘点和下层支撑点",
            "floors": [
                [[11, 11], [12, 12], [13, 13], [21, 21], [22, 22], [31, 31], [32, 32], [33, 33], [41, 41], [44, 44],
                 [42, 42], [71, 71], [72, 72], [73, 73]],
                [[51, 51], [61, 61]]
            ],
            "name_pairs": [
                [["sk_3", "e_3"], ["sk_4", "e_4"]],
                [["sk_5", "e_3"], ["sk_8", "e_4"]]
            ]
        },
        {
            "description": "前层边缘点和上层支撑点",
            "floors": [
                [[11, 11], [12, 12], [13, 13], [21, 21], [22, 22], [31, 31], [32, 32], [33, 33], [41, 41], [44, 44],
                 [42, 42], [71, 71], [72, 72], [73, 73]],
                [[52, 51], [62, 61]]
            ],
            "name_pairs": [
                [["sk_5", "e_1"], ["sk_6", "e_2"]],
                [["sk_1", "e_1"], ["sk_4", "e_2"]]
            ]
        },
        {
            "description": "后层边缘点和上层支撑点",
            "floors": [
                [[11, 11], [12, 12], [13, 13], [21, 21], [22, 22], [31, 31], [32, 32], [33, 33], [41, 41], [44, 44],
                 [42, 42], [71, 71], [72, 72], [73, 73]],
                [[52, 51], [62, 61]]
            ],
            "name_pairs": [
                [["sk_7", "e_3"], ["sk_8", "e_4"]],
                [["sk_5", "e_3"], ["sk_8", "e_4"]]
            ]
        },
        {
            "description": "边缘点前后相连",
            "floors": [
                [[11, 11], [12, 12], [13, 13], [21, 21], [22, 22], [23, 23], [31, 31], [32, 32], [33, 33], [41, 41],
                 [42, 42], [44, 44], [51, 51], [61, 61], [71, 71], [72, 72], [73, 73]],
            ],
            "name_pairs": [
                [["e_4", "e_1"], ["e_3", "e_2"]],
            ]
        },
        {
            "description": "前面下层支撑点水平相连",
            "floors": [
                [[11, 11], [12, 12], [13, 13], [14, 14], [21, 21], [22, 22], [24, 24], [31, 31], [32, 32], [33, 33],
                 [34, 34], [41, 41], [42, 42], [43, 43], [44, 44], [71, 71], [72, 72], [73, 73], [74, 74]],
                [[51, 51], [61, 61]]
            ],
            "name_pairs": [
                [["sk_1", "sk_2"]],
                [["sk_1", "sk_2"], ["sk_3", "sk_4"]]
            ],
            "no_direction": True
        },
        {
            "description": "后面下层支撑点水平相连",
            "floors": [
                [[11, 11], [12, 12], [13, 13], [14, 14], [21, 21], [22, 22], [24, 24], [31, 31], [32, 32], [33, 33],
                 [34, 34], [41, 41], [42, 42], [43, 43], [44, 44], [71, 71], [72, 72], [73, 73], [74, 74]],
                [[51, 51], [61, 61]]
            ],
            "name_pairs": [
                [["sk_4", "sk_3"]],
                [["sk_8", "sk_7"], ["sk_6", "sk_5"]]
            ],
            "no_direction": True
        },
        {
            "description": "前面上层支撑点水平相连",
            "floors": [
                [[11, 11], [12, 12], [13, 13], [14, 14], [21, 21], [22, 22], [24, 24], [31, 31], [32, 32], [33, 33],
                 [34, 34], [41, 41], [42, 42], [43, 43], [44, 44], [71, 71], [72, 72], [73, 73], [74, 74]],
                [[52, 52], [62, 62]]
            ],
            "name_pairs": [
                [["sk_5", "sk_6"]],
                [["sk_1", "sk_2"], ["sk_3", "sk_4"]]
            ],
            "no_direction": True
        },
        {
            "description": "后面上层支撑点水平相连",
            "floors": [
                [[11, 11], [12, 12], [13, 13], [14, 14], [21, 21], [22, 22], [24, 24], [31, 31], [32, 32], [33, 33],
                 [34, 34], [41, 41], [42, 42], [43, 43], [44, 44], [71, 71], [72, 72], [73, 73], [74, 74]],
                [[52, 52], [62, 62]]
            ],
            "name_pairs": [
                [["sk_8", "sk_7"]],
                [["sk_8", "sk_7"], ["sk_6", "sk_5"]]
            ],
            "no_direction": True
        },
        {
            "description": "下层支撑点前后相连",
            "floors": [
                [[11, 11], [12, 12], [13, 13], [14, 14], [21, 21], [22, 22], [24, 24], [31, 31], [32, 32], [33, 33],
                 [34, 34], [41, 41], [42, 42], [43, 43], [44, 44], [71, 71], [72, 72], [73, 73], [74, 74]],
                [[51, 51], [61, 61]]
            ],
            "name_pairs": [
                [["sk_4", "sk_1"], ["sk_3", "sk_2"]],
                [["sk_8", "sk_1"], ["sk_7", "sk_2"], ["sk_6", "sk_3"], ["sk_5", "sk_4"]]
            ]
        },
        {
            "description": "上层支撑点前后相连",
            "floors": [
                [[11, 11], [12, 12], [13, 13], [14, 14], [21, 21], [22, 22], [24, 24], [31, 31], [32, 32], [33, 33],
                 [34, 34], [41, 41], [42, 42], [43, 43], [44, 44], [71, 71], [72, 72], [73, 73], [74, 74]],
                [[52, 52], [62, 62]]
            ],
            "name_pairs": [
                [["sk_8", "sk_5"], ["sk_7", "sk_6"]],
                [["sk_8", "sk_1"], ["sk_7", "sk_2"], ["sk_6", "sk_3"], ["sk_5", "sk_4"]]
            ]
        },
        {
            "description": "前层同层支撑点上下相连",
            "floors": [
                [[11, 11], [12, 12], [13, 13], [14, 14], [21, 21], [22, 22], [24, 24], [31, 31], [32, 32], [33, 33],
                 [34, 34], [41, 41], [42, 42], [43, 43], [44, 44], [71, 71], [72, 72], [73, 73], [74, 74]],
                [[51, 52], [61, 62]]
            ],
            "name_pairs": [
                [["sk_1", "sk_5"], ["sk_2", "sk_6"]],
                [["sk_1", "sk_1"], ["sk_2", "sk_2"], ["sk_3", "sk_3"], ["sk_4", "sk_4"]]
            ]
        },
        {
            "description": "后层同层支撑点上下相连",
            "floors": [
                [[11, 11], [12, 12], [13, 13], [14, 14], [21, 21], [22, 22], [24, 24], [31, 31], [32, 32], [33, 33],
                 [34, 34], [41, 41], [42, 42], [43, 43], [44, 44], [71, 71], [72, 72], [73, 73], [74, 74]],
                [[51, 52], [61, 62]]
            ],
            "name_pairs": [
                [["sk_4", "sk_8"], ["sk_3", "sk_7"]],
                [["sk_8", "sk_8"], ["sk_7", "sk_7"], ["sk_6", "sk_6"], ["sk_5", "sk_5"]]
            ]
        },
        {
            "description": "前层隔层支撑点上下相连",
            "floors": [
                [[11, 12], [12, 13], [13, 14], [21, 22], [22, 24], [31, 32], [32, 33], [33, 34],
                 [41, 42], [42, 43], [71, 72], [72, 73], [73, 74]]
            ],
            "name_pairs": [
                [["sk_5", "sk_1"], ["sk_6", "sk_2"]]
            ]

        },
        {
            "description": "后层隔层支撑点上下相连",
            "floors": [
                [[11, 12], [12, 13], [13, 14], [21, 22], [22, 24], [31, 32], [32, 33], [33, 34],
                 [41, 42], [42, 43], [71, 72], [72, 73], [73, 74]]
            ],
            "name_pairs": [
                [["sk_8", "sk_4"], ["sk_7", "sk_3"]]
            ]
        },
        {
            "description": "前后上隔层支撑点水平相连",
            "floors": [
                [[43, 44]]
            ],
            "name_pairs": [
                [["sk_5", "sk_7"]]
            ]
        }
        ,
        {
            "description": "后后上隔层支撑点水平相连",
            "floors": [
                [[43, 44]]
            ],
            "name_pairs": [
                [["sk_8", "sk_8"]]
            ]
        },
        {
            "description": "后后上隔层支撑点水平相连_酒杯型",
            "floors": [
                [[52, 52], [62, 62]]
            ],
            "name_pairs": [
                [["sk_7", "sk_6"]]
            ],
            "no_direction": True
        },
        {
            "description": "前前上隔层支撑点水平相连",
            "floors": [
                [[52, 52], [62, 62]]
            ],
            "name_pairs": [
                [["sk_2", "sk_3"]]
            ],
            "no_direction": True
        },
        {
            "description": "前前下隔层支撑点水平相连",
            "floors": [
                [[51, 51], [61, 61]]
            ],
            "name_pairs": [
                [["sk_2", "sk_3"]]
            ],
            "no_direction": True
        },
        {
            "description": "后后下隔层支撑点水平相连",
            "floors": [
                [[51, 51], [61, 61]]
            ],
            "name_pairs": [
                [["sk_7", "sk_6"]]
            ],
            "no_direction": True
        },
        {
            "description": "前面非边缘绝缘子点和支撑点相连",
            "floors": [
                [[51, 51], [61, 61]]
            ],
            "name_pairs": [
                [["sk_2", "se_1"], ["sk_3", "se_1"]]
            ]
        },
        {
            "description": "后面非边缘绝缘子点和支撑点相连",
            "floors": [
                [[51, 51], [61, 61]]
            ],
            "name_pairs": [
                [["sk_6", "se_2"], ["sk_7", "se_2"]]
            ]
        },
        {
            "description": "非边缘绝缘子点前后相连",
            "floors": [
                [[51, 51], [61, 61]]
            ],
            "name_pairs": [
                [["se_2", "se_1"], ["se_2", "se_1"]]
            ]
        },
        {
            "description": "前下支撑点与前避雷线点",
            "floors": [
                [[14, 14], [24, 24], [34, 34], [43, 43]]
            ],
            "name_pairs": [
                [["sk_1", "bl_1"], ["sk_2", "bl_2"]]
            ]
        },
        {
            "description": "前上支撑点与前避雷线点",
            "floors": [
                [[14, 14], [24, 24], [34, 34], [43, 43]],
                [[52, 53], [62, 63]],
                [[74, 74]]
            ],
            "name_pairs": [
                [["sk_5", "bl_1"], ["sk_6", "bl_2"]],
                [["sk_1", "bl_1"], ["sk_2", "bl_1"], ["sk_3", "bl_2"], ["sk_4", "bl_2"]],
                [["sk_5", "bl_1"], ["sk_6", "bl_1"]]
            ]
        },
        {
            "description": "后下支撑点与后避雷线点",
            "floors": [
                [[14, 14], [24, 24], [34, 34], [43, 43]]
            ],
            "name_pairs": [
                [["sk_4", "bl_4"], ["sk_3", "bl_3"]]
            ]
        },
        {
            "description": "后上支撑点与后避雷线点",
            "floors": [
                [[14, 14], [24, 24], [34, 34], [43, 43]],
                [[52, 53], [62, 63]],
                [[74, 74]]
            ],
            "name_pairs": [
                [["sk_8", "bl_4"], ["sk_7", "bl_3"]],
                [["sk_8", "bl_4"], ["sk_7", "bl_4"], ["sk_6", "bl_3"], ["sk_5", "bl_3"]],
                [["sk_8", "bl_4"], ["sk_7", "bl_4"]]
            ]
        },
        # {
        #     "description": "酒杯型避雷线点与前支撑点相连",
        #     "floors": [
        #         [[52, 53], [62, 63]]
        #     ],
        #     "name_pairs": [
        #         [["sk_1", "bl_1"], ["sk_2", "bl_1"], ["sk_3", "bl_2"], ["sk_4", "bl_2"]]
        #     ]
        # },
        # {
        #     "description": "酒杯型避雷线点与后支撑点相连",
        #     "floors": [
        #         [[52, 53], [62, 63]]
        #     ],
        #     "name_pairs": [
        #         [["sk_8", "bl_4"], ["sk_7", "bl_4"], ["sk_6", "bl_3"], ["sk_5", "bl_3"]]
        #     ]
        # },
        {
            "description": "避雷线点前后相连",
            "floors": [
                [[14, 14], [24, 24], [34, 34], [43, 43], [53, 53], [63, 63]]
            ],
            "name_pairs": [
                [["bl_4", "bl_1"], ["bl_3", "bl_2"]]
            ]
        },
        {
            "description": "下前特殊支撑点-前边缘点",
            "floors": [
                [[23, 23]]
            ],
            "name_pairs": [
                [["ssk_1", "e_1"], ["ssk_2", "e_2"]]
            ]
        },
        {
            "description": "上前特殊支撑点-前边缘点",
            "floors": [
                [[23, 23]]
            ],
            "name_pairs": [
                [["ssk_5", "e_1"], ["ssk_6", "e_2"]]
            ]
        },
        {
            "description": "下后特殊支撑点-后边缘点",
            "floors": [
                [[23, 23]]
            ],
            "name_pairs": [
                [["ssk_4", "e_4"], ["ssk_3", "e_3"]]
            ]
        },
        {
            "description": "上后特殊支撑点-后边缘点",
            "floors": [
                [[23, 23]]
            ],
            "name_pairs": [
                [["ssk_8", "e_4"], ["ssk_7", "e_3"]]
            ]
        },
        {
            "description": "上层特殊支撑点前后相连",
            "floors": [
                [[23, 23]]
            ],
            "name_pairs": [
                [["ssk_8", "ssk_5"], ["ssk_7", "ssk_6"]]
            ]
        },
        {
            "description": "下层特殊支撑点前后相连",
            "floors": [
                [[23, 23]]
            ],
            "name_pairs": [
                [["ssk_4", "ssk_1"], ["ssk_3", "ssk_2"]]
            ]
        },
        {
            "description": "前面特殊支撑点下上相连",
            "floors": [
                [[23, 23]]
            ],
            "name_pairs": [
                [["ssk_1", "ssk_5"], ["ssk_2", "ssk_6"]]
            ]
        },
        {
            "description": "后面特殊支撑点下上相连",
            "floors": [
                [[23, 23]]
            ],
            "name_pairs": [
                [["ssk_4", "ssk_8"], ["ssk_3", "ssk_7"]]
            ]
        },
    ]

    book = openpyxl.load_workbook(r"new_points.xlsx")
    # sh = book.create_sheet("Sheet7")
    sh = book["Sheet7"]
    sh['A1'] = 'p1'
    sh['B1'] = 'p2'
    sh['C1'] = 'skeleton_type'

    r_num = 2
    for skeleton_type_, skeleton_data in enumerate(temp_dict):
        skeleton_type = skeleton_type_ + 1
        if "floors" in skeleton_data.keys():
            if len(skeleton_data['floors']) != 0:
                for (sk_floors, sk_name_pairs) in zip(skeleton_data['floors'], skeleton_data['name_pairs']):
                    for [pt1_floor, pt2_floor] in sk_floors:
                        for [pt1_name, pt2_name] in sk_name_pairs:
                            sh[f'A{r_num}'] = f"{str(pt1_floor)[0]}_{str(pt1_floor)[1]}_{pt1_name}"
                            sh[f'B{r_num}'] = f"{str(pt2_floor)[0]}_{str(pt2_floor)[1]}_{pt2_name}"
                            sh[f'C{r_num}'] = skeleton_type
                            r_num += 1
        if "no_direction" in skeleton_data.keys():
            print("-----------------------")
            print(skeleton_type)
            print("-----------------------")

    book.save(r"new_points.xlsx")


def get_sk_ids():
    book = openpyxl.load_workbook(r"new_points.xlsx")
    pt_2_type = {}
    rom_num = 2

    sk_kt_channels = [[]] * 37
    sk_channels = []
    sh = book['Sheet6']
    while sh[f'A{rom_num}'].value != None:
        pt_2_type[sh[f'A{rom_num}'].value] = sh[f'B{rom_num}'].value - 1
        rom_num += 1

    sh = book['Sheet7']
    rom_num = 2
    while sh[f'A{rom_num}'].value != None:
        p1_name = sh[f'A{rom_num}'].value
        p2_name = sh[f'B{rom_num}'].value
        sk_type = sh[f'C{rom_num}'].value
        if sk_kt_channels[sk_type - 1] == []:
            sk_kt_channels[sk_type - 1] = [pt_2_type[p1_name], pt_2_type[p2_name]]
        else:
            if not sk_kt_channels[sk_type - 1] == [pt_2_type[p1_name], pt_2_type[p2_name]]:
                print(p1_name, pt_2_type[p1_name], p2_name, pt_2_type[p2_name], sk_type)
                exit()
        rom_num += 1
    for i in range(37):
        sk_channels.append([i * 2, i * 2 + 1])
    print(sk_kt_channels)
    print(sk_channels)
    a = 1


def get_only_one_sample():
    src_dir = r"../data/00_Tower_Dataset"

    json_path = r"../data/coco/annotations/person_keypoints_train2017.json"
    template_data = json.load(open(json_path, "r", encoding="utf-8"), strict=False)

    for input_size in [256, 384, 512, 640, 768, 1024, 1280]:
        img_id = 0
        ann_id = 0
        for dataset in ['train', 'val', 'test']:
            output_data = copy.deepcopy(template_data)
            output_data['categories'] = [{'supercategory': "tower",
                                          'id': 0,
                                          'name': 'tower',
                                          'keypoints': ['edgepoint',
                                                        'skeletonpoints',
                                                        'edgepoint_2',
                                                        'bileipoint',
                                                        'skeletonpoint_2'],
                                          'skeleton': []
                                          }]
            output_data['images'] = []
            output_data['annotations'] = []

            json_path = r"{}/{}/anns/tower_info_{}.json".format(src_dir, input_size, dataset)
            data = json.load(open(json_path, "r", encoding="utf-8"), strict=False)

            num = 0

            for (img_name, temp_dict) in data.items():
                width = temp_dict['width']
                height = temp_dict['height']
                keypoint_list = [width // 2, height // 2, 2] * 5
                [x1, y1, x2, y2] = temp_dict['bbox']

                img_dict = {
                    "license": 1,
                    "file_name": img_name,
                    "coco_url": "",
                    "height": height,
                    "width": width,
                    "date_captured": "",
                    "flickr_url": "",
                    "id": img_id
                }

                ann_dict = {
                    "segmentation": [[x1, y1, x2, y1, x2, y2, x1, y2]],
                    "true_points_1": np.array(temp_dict["new_keypoints_1"])[:, 1:].astype(np.int32).tolist(),
                    "true_skeletons_1": temp_dict["new_skeletons_1"],
                    "true_points_2": np.array(temp_dict["new_keypoints_2"])[:, 1:].astype(np.int32).tolist(),
                    "true_skeletons_2": temp_dict["new_skeletons_2"],
                    "true_points_3": np.array(temp_dict["new_keypoints_3"])[:, 1:].astype(np.int32).tolist(),
                    "true_skeletons_3": temp_dict["new_skeletons_3"],
                    "num_keypoints": 5,
                    "area": int((x2 - x1) * (y2 - y1)),
                    "iscrowd": 0,
                    "keypoints": keypoint_list,
                    "image_id": img_id,
                    "bbox": [x1, y1, x2 - x1, y2 - y1],
                    "category_id": 1,
                    "id": ann_id
                }
                output_data['images'].append(img_dict)
                output_data['annotations'].append(ann_dict)

                img_id += 1
                ann_id += 1
                num += 1
                if num > 2:
                    break

            out_json_path = r"{}/{}/anns/tower_keypoints_{}_2.json".format(src_dir, input_size, dataset)
            f = open(out_json_path, "w", encoding="utf-8")
            f.write(json.dumps(output_data, ensure_ascii=False, indent=4))
            f.close()


if __name__ == '__main__':
    scale = 8
    sigma = 2
    half_width = 1
    out_img_size = 512
    num_point_type = 14
    num_skeleton_type = 39
    ue_src_dir = r""
    bgr_colors = []
    hue_steps = num_skeleton_type  # 均分的色调数量
    saturation = 1.0  # 最大饱和度
    value = 1.0  # 最大明度
    for i in range(hue_steps):
        hue = i / hue_steps  # 计算色调
        rgb_color = colorsys.hsv_to_rgb(hue, saturation, value)  # 转换为RGB颜色
        bgr_color = tuple([round(x * 255) for x in reversed(rgb_color)])  # 转换为BGR颜色
        bgr_colors.append(bgr_color)
    sk_colores = bgr_colors

    bgr_colors = []
    hue_steps = num_point_type  # 均分的色调数量
    saturation = 1.0  # 最大饱和度
    value = 1.0  # 最大明度
    for i in range(hue_steps):
        hue = i / hue_steps  # 计算色调
        rgb_color = colorsys.hsv_to_rgb(hue, saturation, value)  # 转换为RGB颜色
        bgr_color = tuple([round(x * 255) for x in reversed(rgb_color)])  # 转换为BGR颜色
        bgr_colors.append(bgr_color)
    pt_colors = bgr_colors

    # get_complex_skeletons()
    # main()
    # step2()
    get_finale_ann()
    # get_only_one_sample()
    # get_sample()
    # get_sk_ids()
